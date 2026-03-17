"""
Generate Excel and PDF from extracted order data.
Fixes:
  - Excel: column widths, RTL alignment, bold headers, proper totals row
  - PDF: RTL text direction, proper Hebrew rendering, page header with all fields
"""
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

# ── Hebrew font registration ────────────────────────────────────────────────

HEBREW_FONT_NAME = "Helvetica"
_font_registered = False


def _register_hebrew_font() -> str:
    global HEBREW_FONT_NAME, _font_registered
    if _font_registered:
        return HEBREW_FONT_NAME
    try:
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        import os

        win_fonts = os.environ.get("SystemRoot", "C:\\Windows") + "\\Fonts"
        candidates = [
            ("TahomaHebrew",   Path(win_fonts) / "tahoma.ttf"),
            ("ArialHebrew",    Path(win_fonts) / "arial.ttf"),
            ("ArialUniHebrew", Path(win_fonts) / "arialuni.ttf"),
            ("DavidHebrew",    Path(win_fonts) / "david.ttf"),
            # Linux / Mac
            ("DejaVuHebrew",   Path("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf")),
            ("LiberationHebrew", Path("/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf")),
            ("MacArialHebrew", Path("/System/Library/Fonts/Supplemental/Arial.ttf")),
        ]
        for name, path in candidates:
            if path.exists():
                try:
                    pdfmetrics.registerFont(TTFont(name, str(path)))
                    HEBREW_FONT_NAME = name
                    _font_registered = True
                    return HEBREW_FONT_NAME
                except Exception:
                    continue
    except Exception:
        pass
    _font_registered = True
    return HEBREW_FONT_NAME


# ── Helpers ──────────────────────────────────────────────────────────────────

def panel_name(panel: dict, index: int) -> str:
    """Format panel name: S-XXX (without quantity suffix in name column)."""
    pid = panel.get("panel_id") or str(index + 1)
    if isinstance(pid, str) and pid.isdigit():
        return f"S-{100 + int(pid)}"
    return str(pid)


def get_width(panel: dict) -> float:
    """Get effective width_mm or derive from profile_dimensions."""
    w = panel.get("width_mm")
    if w is not None:
        try:
            return float(w)
        except (TypeError, ValueError):
            pass
    dims = panel.get("profile_dimensions")
    if dims:
        try:
            return float(sum(float(d) for d in dims[::2]))  # sum of horizontal segments
        except (TypeError, ValueError):
            pass
    return 0.0


# ── Excel ─────────────────────────────────────────────────────────────────────

FONT_HEB   = Font(name="Arial", size=11)
FONT_BOLD  = Font(name="Arial", size=11, bold=True)
FONT_TITLE = Font(name="Arial", size=13, bold=True)
FONT_SMALL = Font(name="Arial", size=9)

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_RIGHT  = Alignment(horizontal="right",  vertical="center", wrap_text=True)
ALIGN_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

FILL_HEADER = PatternFill("solid", fgColor="1F4E79")   # dark blue
FILL_TOTAL  = PatternFill("solid", fgColor="D9E1F2")   # light blue
FILL_ALT    = PatternFill("solid", fgColor="F5F7FA")   # very light grey (alternating rows)

BORDER_THIN = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)


def _set(ws, row, col, value, font=None, align=None, fill=None, border=None):
    cell = ws.cell(row=row, column=col, value=value)
    if font:   cell.font   = font
    if align:  cell.alignment = align
    if fill:   cell.fill   = fill
    if border: cell.border = border
    return cell


def _header_block(ws, header: dict, project_id: str) -> None:
    """Write the 3-row header block."""
    # Row 1
    _set(ws, 1, 1, header.get("material") or "אלומיניום", FONT_BOLD, ALIGN_RIGHT)
    _set(ws, 1, 2, 'ח"ג:',   FONT_HEB, ALIGN_RIGHT)
    _set(ws, 1, 3, header.get("date") or "",              FONT_HEB, ALIGN_CENTER)
    _set(ws, 1, 6, "תאריך:", FONT_HEB, ALIGN_RIGHT)
    _set(ws, 1, 7, header.get("client_name") or "",       FONT_HEB, ALIGN_RIGHT)

    # Row 2
    thick = header.get("thickness_mm") or 2
    _set(ws, 2, 1, f'{thick} מ"מ',                        FONT_HEB, ALIGN_RIGHT)
    _set(ws, 2, 2, "עובי:",  FONT_HEB, ALIGN_RIGHT)
    _set(ws, 2, 3, header.get("project_id") or project_id, FONT_BOLD, ALIGN_CENTER)
    _set(ws, 2, 6, "הזמנה:", FONT_HEB, ALIGN_RIGHT)
    _set(ws, 2, 7, header.get("project_name") or "",      FONT_HEB, ALIGN_RIGHT)

    # Row 3
    _set(ws, 3, 1, header.get("color_ral") or "9011",    FONT_HEB, ALIGN_RIGHT)
    _set(ws, 3, 2, "צבע:",   FONT_HEB, ALIGN_RIGHT)
    _set(ws, 3, 6, "הערות:", FONT_HEB, ALIGN_RIGHT)
    _set(ws, 3, 7, header.get("order_number") or "",      FONT_HEB, ALIGN_RIGHT)


COL_HEADERS = ["#", "שם", 'אורך\nמ"מ', 'רוחב\nמ"מ', "כמות", "לסובב", "חומר", 'שטח\nמ"ר', "הערות"]
COL_WIDTHS  = [5,   14,   10,            10,            7,      8,       14,     10,          20]


def create_excel(data: dict, output_path: Path) -> None:
    """Create Excel order file with proper formatting and RTL."""
    header = data.get("header", {})
    panels = data.get("panels", [])

    wb = Workbook()
    ws = wb.active
    ws.title = "הזמנה"
    ws.sheet_view.rightToLeft = True   # ← RTL sheet

    _header_block(ws, header, header.get("project_id", ""))

    # Empty row 4, column headers in row 5
    for col, (label, width) in enumerate(zip(COL_HEADERS, COL_WIDTHS), start=1):
        cell = _set(ws, 5, col, label,
                    Font(name="Arial", size=11, bold=True, color="FFFFFF"),
                    ALIGN_CENTER, FILL_HEADER, BORDER_THIN)
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[5].height = 30

    total_area = 0.0
    for i, panel in enumerate(panels):
        row = 6 + i
        length = float(panel.get("length_mm") or 0)
        width  = get_width(panel)
        qty    = int(panel.get("quantity") or 1)
        area   = (length * width * qty) / 1_000_000 if width else 0
        total_area += area
        thick  = header.get("thickness_mm", 2)

        fill = FILL_ALT if i % 2 == 1 else None

        _set(ws, row, 1, i + 1,                           FONT_HEB,  ALIGN_CENTER, fill, BORDER_THIN)
        _set(ws, row, 2, panel_name(panel, i),            FONT_HEB,  ALIGN_RIGHT,  fill, BORDER_THIN)
        _set(ws, row, 3, length,                          FONT_HEB,  ALIGN_CENTER, fill, BORDER_THIN)
        _set(ws, row, 4, width if width else None,        FONT_HEB,  ALIGN_CENTER, fill, BORDER_THIN)
        _set(ws, row, 5, qty,                             FONT_HEB,  ALIGN_CENTER, fill, BORDER_THIN)
        _set(ws, row, 6, panel.get("turn") or "N",        FONT_HEB,  ALIGN_CENTER, fill, BORDER_THIN)
        _set(ws, row, 7, f"Al {thick} mm",                FONT_HEB,  ALIGN_CENTER, fill, BORDER_THIN)
        _set(ws, row, 8, round(area, 3),                  FONT_HEB,  ALIGN_CENTER, fill, BORDER_THIN)
        _set(ws, row, 9, panel.get("notes") or "",        FONT_HEB,  ALIGN_RIGHT,  fill, BORDER_THIN)

    # Totals row
    tot_row = 6 + len(panels)
    _set(ws, tot_row, 1, len(panels),         FONT_BOLD, ALIGN_CENTER, FILL_TOTAL, BORDER_THIN)
    _set(ws, tot_row, 2, "סה\"כ",             FONT_BOLD, ALIGN_RIGHT,  FILL_TOTAL, BORDER_THIN)
    for col in range(3, 8):
        _set(ws, tot_row, col, None,          FONT_HEB,  ALIGN_CENTER, FILL_TOTAL, BORDER_THIN)
    _set(ws, tot_row, 8, round(total_area, 3), FONT_BOLD, ALIGN_CENTER, FILL_TOTAL, BORDER_THIN)
    _set(ws, tot_row, 9, "",                  FONT_HEB,  ALIGN_RIGHT,  FILL_TOTAL, BORDER_THIN)

    # Freeze panes below header + column headers
    ws.freeze_panes = "A6"

    wb.save(output_path)


# ── PDF ───────────────────────────────────────────────────────────────────────

def _visual(text: str) -> str:
    """Convert logical-order Hebrew to visual order for PDF rendering."""
    try:
        from bidi.algorithm import get_display
        return get_display(text)
    except ImportError:
        return text


def create_pdf(data: dict, output_path: Path) -> None:
    """Create PDF order summary with proper Hebrew/RTL rendering."""
    header = data.get("header", {})
    panels = data.get("panels", [])
    font_name = _register_hebrew_font()

    doc = SimpleDocTemplate(
        str(output_path),
        pagesize=A4,
        rightMargin=15 * mm,
        leftMargin=15 * mm,
        topMargin=15 * mm,
        bottomMargin=15 * mm,
    )

    styles = getSampleStyleSheet()
    style_title = ParagraphStyle(
        "HebrewTitle",
        parent=styles["Normal"],
        fontName=font_name,
        fontSize=14,
        leading=18,
        alignment=2,  # right
        spaceAfter=4,
    )
    style_sub = ParagraphStyle(
        "HebrewSub",
        parent=styles["Normal"],
        fontName=font_name,
        fontSize=10,
        leading=14,
        alignment=2,
        spaceAfter=2,
    )
    style_small = ParagraphStyle(
        "HebrewSmall",
        parent=styles["Normal"],
        fontName=font_name,
        fontSize=9,
        leading=12,
        alignment=2,
    )

    story = []

    # ── Page title block ──
    thick = header.get("thickness_mm") or 2
    ral   = header.get("color_ral") or "9011"
    mat   = header.get("material") or "אלומיניום"
    title = f'{mat} {thick} מ"מ – RAL {ral}'
    story.append(Paragraph(_visual(title), style_title))

    proj_id   = header.get("project_id") or ""
    proj_name = header.get("project_name") or ""
    client    = header.get("client_name") or ""
    date      = header.get("date") or ""
    order_num = header.get("order_number") or ""

    if proj_id or proj_name:
        story.append(Paragraph(_visual(f"פרויקט: {proj_id}  {proj_name}"), style_sub))
    if client:
        story.append(Paragraph(_visual(f"לקוח: {client}"), style_sub))
    if date or order_num:
        story.append(Paragraph(_visual(f"תאריך: {date}   מס' הזמנה: {order_num}"), style_sub))

    story.append(Spacer(1, 5 * mm))

    # ── Table ──
    # Headers (RTL)
    col_headers = [_visual("הערות"), _visual('שטח מ"ר'), _visual("חומר"), _visual("כמות"),
                   _visual('רוחב מ"מ'), _visual('אורך מ"מ'), _visual("שם"), "#"]
    table_data = [col_headers]

    total_area = 0.0
    for i, panel in enumerate(panels):
        length = float(panel.get("length_mm") or 0)
        width  = get_width(panel)
        qty    = int(panel.get("quantity") or 1)
        area   = (length * width * qty) / 1_000_000 if width else 0
        total_area += area

        notes = panel.get("notes") or ""
        row = [
            _visual(notes) if notes else "",
            f"{area:.3f}",
            f"Al {thick} mm",
            str(qty),
            str(int(width)) if width else "",
            str(int(length)),
            _visual(panel_name(panel, i)),
            str(i + 1),
        ]
        table_data.append(row)

    # Totals row
    table_data.append([
        "",
        f"{total_area:.3f}",
        "",
        str(len(panels)),
        "",
        "",
        _visual('סה"כ'),
        "",
    ])

    # Column widths (page is ~180mm usable) – RTL order
    col_widths = [35*mm, 18*mm, 22*mm, 14*mm, 18*mm, 18*mm, 28*mm, 10*mm]

    t = Table(table_data, colWidths=col_widths, repeatRows=1)

    dark_blue  = colors.HexColor("#1F4E79")
    light_blue = colors.HexColor("#D9E1F2")
    alt_grey   = colors.HexColor("#F5F7FA")

    style_cmds = [
        # Header row
        ("BACKGROUND",  (0, 0), (-1, 0),  dark_blue),
        ("TEXTCOLOR",   (0, 0), (-1, 0),  colors.white),
        ("FONTNAME",    (0, 0), (-1, 0),  font_name),
        ("FONTSIZE",    (0, 0), (-1, 0),  9),
        ("ALIGN",       (0, 0), (-1, 0),  "CENTER"),
        ("VALIGN",      (0, 0), (-1, 0),  "MIDDLE"),
        ("ROWBACKGROUND", (0, 1), (-1, -2), [colors.white, alt_grey]),
        # Totals row
        ("BACKGROUND",  (0, -1), (-1, -1), light_blue),
        ("FONTNAME",    (0, -1), (-1, -1), font_name),
        ("FONTSIZE",    (0, -1), (-1, -1), 9),
        # All cells
        ("FONTNAME",    (0, 1), (-1, -1), font_name),
        ("FONTSIZE",    (0, 1), (-1, -1), 9),
        ("ALIGN",       (0, 1), (-1, -1), "CENTER"),
        ("VALIGN",      (0, 0), (-1, -1), "MIDDLE"),
        ("GRID",        (0, 0), (-1, -1), 0.4, colors.HexColor("#AAAAAA")),
        ("ROWHEIGHT",   (0, 0), (-1, -1), 16),
    ]
    t.setStyle(TableStyle(style_cmds))
    story.append(t)

    doc.build(story)


# ── Main ──────────────────────────────────────────────────────────────────────

def generate_output(data: dict, project_id: str, output_dir: Path) -> dict:
    """Generate Excel, PDF, and DXF. Returns dict of paths."""
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    if "header" not in data:
        data["header"] = {}
    data["header"]["project_id"] = data["header"].get("project_id") or project_id

    excel_path = output_dir / f"______{project_id}-הזמנה.xlsx"
    pdf_path   = output_dir / f"______{project_id}-הזמנה.pdf"

    create_excel(data, excel_path)
    create_pdf(data, pdf_path)

    result = {"excel": str(excel_path), "pdf": str(pdf_path)}

    # DXF files
    try:
        from .dxf_output import create_dxf_files
        dxf_paths = create_dxf_files(data, project_id, output_dir)
        if dxf_paths:
            result["dxf"]     = [str(p) for p in dxf_paths]
            result["dxf_zip"] = str(output_dir / f"DWG-{project_id}.zip")
            result["pdf_zip"] = str(output_dir / f"PDF-{project_id}.zip")
    except ImportError:
        pass

    return result


if __name__ == "__main__":
    import sys
    if len(sys.argv) < 2:
        print("Usage: python -m src.output <json_path> [project_id]")
        sys.exit(1)
    json_path  = sys.argv[1]
    project_id = sys.argv[2] if len(sys.argv) > 2 else "MR277-3"
    with open(json_path, encoding="utf-8") as f:
        data = json.load(f)
    out_dir = Path("output") / project_id
    out_dir.mkdir(parents=True, exist_ok=True)
    result = generate_output(data, project_id, out_dir)
    print("Excel:", result["excel"])
    print("PDF:  ", result["pdf"])
